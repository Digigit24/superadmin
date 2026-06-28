from rest_framework import viewsets, status, permissions
from rest_framework.decorators import action, api_view, permission_classes
from rest_framework.response import Response
from rest_framework_simplejwt.tokens import RefreshToken
from django.contrib.auth import authenticate
from django.http import HttpResponse
from apps.accounts.models import CustomUser, Role, IntegrationToken
from apps.accounts.serializers import (
    UserSerializer, UserCreateSerializer, RoleSerializer, IntegrationTokenSerializer,
    RegisterSerializer, LoginSerializer, ChangePasswordSerializer
)
from apps.accounts.services import get_tokens_for_user
from apps.common.permissions import IsSuperAdmin, IsTenantAdmin, IsTenantMember
from apps.common.constants import PERMISSION_SCHEMA
from apps.common.logger import get_logger
import io
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import json

logger = get_logger(__name__)


@api_view(['POST'])
@permission_classes([permissions.AllowAny])
def register_view(request):
    serializer = RegisterSerializer(data=request.data)
    if serializer.is_valid():
        result = serializer.save()
        user = result['user']
        logger.info(f'New user registered: {user.email}')
        tokens = get_tokens_for_user(user)
        return Response({
            'message': 'Registration successful',
            'user': UserSerializer(user).data,
            'tokens': tokens
        }, status=status.HTTP_201_CREATED)
    logger.warning(f'Registration failed: {serializer.errors}')
    return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


@api_view(['POST'])
@permission_classes([permissions.AllowAny])
def login_view(request):
    serializer = LoginSerializer(data=request.data)
    if serializer.is_valid():
        email = serializer.validated_data['email']
        user = authenticate(
            request,
            username=email,
            password=serializer.validated_data['password']
        )

        if user and user.is_active:
            logger.info(f'User logged in successfully: {user.email}, tenant: {user.tenant.slug if user.tenant else "No tenant"}')
            try:
                tokens = get_tokens_for_user(user)
                logger.debug(f'JWT tokens generated for user: {user.email}')
                return Response({
                    'message': 'Login successful',
                    'user': UserSerializer(user).data,
                    'tokens': tokens
                })
            except Exception as e:
                logger.error(f'Error generating tokens for user {user.email}: {str(e)}', exc_info=True)
                return Response({'error': 'Token generation failed'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

        logger.warning(f'Failed login attempt for email: {email}')
        return Response({'error': 'Invalid credentials'}, status=status.HTTP_401_UNAUTHORIZED)

    logger.warning(f'Invalid login request data: {serializer.errors}')
    return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


@api_view(['POST'])
def logout_view(request):
    try:
        refresh_token = request.data.get('refresh_token')
        token = RefreshToken(refresh_token)
        token.blacklist()
        logger.info(f'User logged out: {request.user.email}')
        return Response({'message': 'Logout successful'})
    except Exception as e:
        logger.warning(f'Logout failed for user {request.user}: {str(e)}')
        return Response({'error': 'Invalid token'}, status=status.HTTP_400_BAD_REQUEST)


@api_view(['POST'])
def change_password_view(request):
    serializer = ChangePasswordSerializer(data=request.data)
    if serializer.is_valid():
        user = request.user
        if not user.check_password(serializer.validated_data['old_password']):
            return Response({'error': 'Current password is incorrect'}, status=status.HTTP_400_BAD_REQUEST)
        
        user.set_password(serializer.validated_data['new_password'])
        user.save()
        return Response({'message': 'Password changed successfully'})
    return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


class UserViewSet(viewsets.ModelViewSet):
    queryset = CustomUser.objects.all()
    
    def get_serializer_class(self):
        if self.action == 'create':
            return UserCreateSerializer
        return UserSerializer
    
    def get_queryset(self):
        user = self.request.user

        # Check for x-tenant-id header for tenant filtering
        x_tenant_id = self.request.headers.get('x-tenant-id') or self.request.headers.get('X-Tenant-Id')

        # If x-tenant-id header is present, filter by that tenant
        if x_tenant_id:
            # Validate access: super admins can access any tenant, others only their own
            if not user.is_super_admin:
                if not user.tenant or str(user.tenant.id) != str(x_tenant_id):
                    logger.warning(f'User {user.email} attempted to access users from different tenant')
                    return CustomUser.objects.none()

            logger.info(f'Filtering users by x-tenant-id header: {x_tenant_id}')
            return CustomUser.objects.filter(tenant=x_tenant_id)

        # Default behavior when no header is present
        if user.is_super_admin:
            return CustomUser.objects.all()
        elif user.tenant:
            return CustomUser.objects.filter(tenant=user.tenant)
        return CustomUser.objects.none()
    
    def get_permissions(self):
        if self.action in ['me', 'update_me', 'create']:
            return [permissions.IsAuthenticated()]
        return [IsTenantAdmin()]
    
    def perform_create(self, serializer):
        user = self.request.user
        # Only set tenant from logged-in user if tenant is not provided in request
        if not user.is_super_admin and 'tenant' not in serializer.validated_data:
            serializer.save(tenant=user.tenant)
        else:
            serializer.save()

    def create(self, request, *args, **kwargs):
        logger.info(f'Create user request from: {request.user.email if request.user.is_authenticated else "Anonymous"}')
        logger.debug(f'Request data: {request.data}')

        # Extract tenant_id from x-tenant-id header if present
        x_tenant_id = request.headers.get('x-tenant-id') or request.headers.get('X-Tenant-Id')

        # Prepare mutable request data
        request_data = request.data.copy()

        # If x-tenant-id header is present and tenant is not in request body, use the header value
        if x_tenant_id and 'tenant' not in request_data:
            request_data['tenant'] = x_tenant_id
            logger.info(f'Tenant ID extracted from x-tenant-id header: {x_tenant_id}')

        # Security check: non-super-admins can only create users in their own tenant
        if not request.user.is_super_admin:
            tenant_id = request_data.get('tenant')
            if tenant_id:
                # Validate tenant exists and user has access
                if not request.user.tenant:
                    logger.error(f'User {request.user.email} has no tenant assigned')
                    return Response(
                        {'error': 'Your account is not associated with a tenant'},
                        status=status.HTTP_400_BAD_REQUEST
                    )

                if str(request.user.tenant.id) != str(tenant_id):
                    logger.warning(f'User {request.user.email} attempted to create user in different tenant')
                    return Response(
                        {'error': 'You can only create users in your own tenant'},
                        status=status.HTTP_403_FORBIDDEN
                    )

        try:
            serializer = self.get_serializer(data=request_data)
            serializer.is_valid(raise_exception=True)
            self.perform_create(serializer)
            user = serializer.instance

            if not user or not user.id:
                logger.error('User object created but has no ID')
                return Response(
                    {'error': 'User creation failed - no ID generated'},
                    status=status.HTTP_500_INTERNAL_SERVER_ERROR
                )

            logger.info(f'User created successfully: {user.email} (ID: {user.id}) in tenant: {user.tenant.slug if user.tenant else "No tenant"}')
            response_data = UserSerializer(user).data
            # Use UserSerializer for the response (includes id and all fields)
           
            

            # Ensure id is in the response
            if 'id' not in response_data:
                logger.error(f'Serializer did not include id field. Data: {response_data}')
                response_data['id'] = str(user.id)

            logger.debug(f'Response data: {response_data}')

            headers = self.get_success_headers(response_data)
            return Response(response_data, status=status.HTTP_201_CREATED, headers=headers)

        except Exception as e:
            logger.error(f'Error creating user: {str(e)}', exc_info=True)
            return Response(
                {'error': f'Failed to create user: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
    
    
    @action(detail=False, methods=['get'])
    def me(self, request):
        serializer = self.get_serializer(request.user)
        return Response(serializer.data)
    
    @action(detail=False, methods=['put', 'patch'])
    def update_me(self, request):
        serializer = self.get_serializer(request.user, data=request.data, partial=True)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
    
    @action(detail=True, methods=['post'])
    def assign_roles(self, request, pk=None):
        user = self.get_object()
        role_ids = request.data.get('role_ids', [])
        roles = Role.objects.filter(id__in=role_ids, tenant=user.tenant)
        user.roles.set(roles)
        return Response({'message': 'Roles assigned successfully'})
    
    @action(detail=True, methods=['delete'])
    def remove_role(self, request, pk=None):
        user = self.get_object()
        role_id = request.data.get('role_id')
        user.roles.remove(role_id)
        return Response({'message': 'Role removed successfully'})

    @action(detail=True, methods=['post'])
    def reset_password(self, request, pk=None):
        """Admin resets a user's password without requiring the old one."""
        user = self.get_object()
        new_password = request.data.get('new_password')
        if not new_password:
            return Response(
                {'error': 'new_password is required'},
                status=status.HTTP_400_BAD_REQUEST
            )
        if len(new_password) < 8:
            return Response(
                {'error': 'Password must be at least 8 characters'},
                status=status.HTTP_400_BAD_REQUEST
            )
        user.set_password(new_password)
        user.save(update_fields=['password'])
        logger.info(f'Password reset for user {user.email} by admin {request.user.email}')
        return Response({'message': 'Password reset successfully'})

    @action(detail=True, methods=['post'])
    def activate(self, request, pk=None):
        """Activate a user account (set is_active=True)."""
        user = self.get_object()
        if user.is_active:
            return Response({'message': 'User is already active'})
        user.is_active = True
        user.save(update_fields=['is_active'])
        logger.info(f'User {user.email} activated by {request.user.email}')
        return Response({'message': 'User activated successfully'})

    @action(detail=True, methods=['post'])
    def deactivate(self, request, pk=None):
        """Deactivate a user account (set is_active=False)."""
        user = self.get_object()
        if not user.is_active:
            return Response({'message': 'User is already inactive'})
        user.is_active = False
        user.save(update_fields=['is_active'])
        logger.info(f'User {user.email} deactivated by {request.user.email}')
        return Response({'message': 'User deactivated successfully'})


    @action(detail=False, methods=['get'])
    def export(self, request):
        """Export all users for this tenant as an Excel file."""
        qs = self.get_queryset()
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Users"

        headers = ['ID', 'Email', 'First Name', 'Last Name', 'Phone', 'Roles',
                   'Timezone', 'Status', 'Joined']
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill("solid", fgColor="1a1a2e")

        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')

        for row_idx, user in enumerate(qs, 2):
            role_names = ", ".join(r.name for r in user.roles.all())
            ws.append([
                str(user.id),
                user.email,
                user.first_name or '',
                user.last_name or '',
                user.phone or '',
                role_names,
                user.timezone or '',
                'Active' if user.is_active else 'Inactive',
                user.date_joined.strftime('%Y-%m-%d') if user.date_joined else '',
            ])

        # Auto-size columns
        for col in ws.columns:
            max_len = max((len(str(cell.value or '')) for cell in col), default=10)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 50)

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        logger.info(f'User export by {request.user.email}: {qs.count()} rows')
        response = HttpResponse(
            buffer.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = 'attachment; filename="users_export.xlsx"'
        return response

    @action(detail=False, methods=['post'])
    def import_users(self, request):
        """Bulk-import users from an uploaded Excel file.

        Expected columns (header row): Email, First Name, Last Name, Phone, Timezone, Password
        Returns a summary: { created, skipped, errors }
        """
        file = request.FILES.get('file')
        if not file:
            return Response({'error': 'No file uploaded. Send as multipart/form-data with key "file"'},
                            status=status.HTTP_400_BAD_REQUEST)

        try:
            wb = openpyxl.load_workbook(file, read_only=True, data_only=True)
        except Exception as e:
            return Response({'error': f'Cannot read Excel file: {e}'}, status=status.HTTP_400_BAD_REQUEST)

        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return Response({'error': 'File is empty'}, status=status.HTTP_400_BAD_REQUEST)

        # Normalise header row
        header = [str(h).strip().lower() if h else '' for h in rows[0]]
        col = {name: idx for idx, name in enumerate(header)}

        def get(row, key, default=''):
            idx = col.get(key)
            return str(row[idx]).strip() if idx is not None and row[idx] is not None else default

        created, skipped, errors = 0, 0, []
        tenant = request.user.tenant if not request.user.is_super_admin else None

        for row_num, row in enumerate(rows[1:], start=2):
            email = get(row, 'email')
            if not email:
                errors.append({'row': row_num, 'error': 'Missing email'})
                continue

            if CustomUser.objects.filter(email=email).exists():
                skipped += 1
                continue

            try:
                password = get(row, 'password') or 'Celiyo@123'  # default temp password
                user = CustomUser.objects.create_user(
                    email=email,
                    password=password,
                    first_name=get(row, 'first name') or get(row, 'first_name'),
                    last_name=get(row, 'last name') or get(row, 'last_name'),
                    phone=get(row, 'phone') or None,
                    timezone=get(row, 'timezone') or 'UTC',
                    tenant=tenant,
                    is_active=True,
                )
                created += 1
            except Exception as e:
                errors.append({'row': row_num, 'email': email, 'error': str(e)})

        logger.info(f'User import by {request.user.email}: {created} created, {skipped} skipped, {len(errors)} errors')
        return Response({
            'created': created,
            'skipped': skipped,
            'errors': errors,
            'message': f'{created} users created, {skipped} skipped (already exist), {len(errors)} failed.',
        }, status=status.HTTP_200_OK)


class RoleViewSet(viewsets.ModelViewSet):
    serializer_class = RoleSerializer
    permission_classes = [IsTenantAdmin]
    
    def get_queryset(self):
        user = self.request.user
        tenant_id = (
            self.request.headers.get('x-tenant-id') or
            self.request.headers.get('X-Tenant-Id') or
            self.request.query_params.get('tenant')
        )

        if tenant_id:
            if not user.is_super_admin:
                if not user.tenant or str(user.tenant.id) != str(tenant_id):
                    logger.warning(f'User {user.email} attempted to access roles from different tenant')
                    return Role.objects.none()
            return Role.objects.filter(tenant=tenant_id)

        if user.is_super_admin:
            return Role.objects.all()
        elif user.tenant:
            return Role.objects.filter(tenant=user.tenant)
        return Role.objects.none()
    
    def perform_create(self, serializer):
        user = self.request.user
        tenant_id = (
            self.request.headers.get('x-tenant-id') or
            self.request.headers.get('X-Tenant-Id') or
            self.request.data.get('tenant')
        )

        if user.is_super_admin and tenant_id:
            serializer.save(tenant_id=tenant_id, created_by=user)
        else:
            serializer.save(tenant=user.tenant, created_by=user)
    
    @action(detail=False, methods=['get'], permission_classes=[permissions.IsAuthenticated])
    def permissions_schema(self, request):
        return Response(PERMISSION_SCHEMA)
    
    @action(detail=True, methods=['get'])
    def members(self, request, pk=None):
        role = self.get_object()
        users = role.users.all()
        return Response(UserSerializer(users, many=True).data)

    @action(detail=False, methods=['get'])
    def export(self, request):
        """Export all roles for this tenant as an Excel file."""
        qs = self.get_queryset()
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Roles"

        headers = ['ID', 'Name', 'Description', 'Status', 'Members', 'Created By', 'Created At']
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill("solid", fgColor="1a1a2e")

        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')

        for role in qs:
            ws.append([
                str(role.id),
                role.name,
                role.description or '',
                'Active' if role.is_active else 'Inactive',
                role.users.count(),
                role.created_by.email if role.created_by else '',
                role.created_at.strftime('%Y-%m-%d') if role.created_at else '',
            ])

        for col in ws.columns:
            max_len = max((len(str(cell.value or '')) for cell in col), default=10)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 50)

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        response = HttpResponse(
            buffer.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = 'attachment; filename="roles_export.xlsx"'
        return response


class IntegrationTokenViewSet(viewsets.ModelViewSet):
    serializer_class = IntegrationTokenSerializer
    permission_classes = [IsTenantAdmin]

    def get_queryset(self):
        user = self.request.user
        if user.is_super_admin:
            return IntegrationToken.objects.all()
        if user.tenant:
            return IntegrationToken.objects.filter(tenant=user.tenant)
        return IntegrationToken.objects.none()

    def get_serializer_context(self):
        context = super().get_serializer_context()
        context['tenant'] = self.request.user.tenant
        return context

    def perform_create(self, serializer):
        serializer.save(tenant=self.request.user.tenant, created_by=self.request.user)

    def create(self, request, *args, **kwargs):
        if not request.user.tenant:
            return Response(
                {'error': 'Your account is not associated with a tenant'},
                status=status.HTTP_400_BAD_REQUEST
            )

        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        self.perform_create(serializer)
        token = serializer.instance.generate_jwt()

        data = serializer.data
        data['system_token'] = token
        headers = self.get_success_headers(data)
        return Response(data, status=status.HTTP_201_CREATED, headers=headers)

    @action(detail=True, methods=['post'])
    def rotate(self, request, pk=None):
        integration_token = self.get_object()
        integration_token.rotate()
        serializer = self.get_serializer(integration_token)
        data = serializer.data
        data['system_token'] = integration_token.generate_jwt()
        return Response(data)

    @action(detail=True, methods=['post'])
    def revoke(self, request, pk=None):
        integration_token = self.get_object()
        integration_token.revoke()
        return Response({'message': 'Integration token revoked successfully'})

    @action(detail=True, methods=['post'])
    def reactivate(self, request, pk=None):
        integration_token = self.get_object()
        integration_token.is_active = True
        integration_token.revoked_at = None
        integration_token.save(update_fields=['is_active', 'revoked_at', 'updated_at'])
        serializer = self.get_serializer(integration_token)
        data = serializer.data
        data['system_token'] = integration_token.generate_jwt()
        return Response(data)
